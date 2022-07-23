from datetime import datetime, timedelta
from io import BytesIO

from django.test import TestCase
from django.urls import reverse
from django.utils import timezone
from openpyxl import load_workbook

from signup.faculty.spreadsheets import generate_spreadsheet
from signup.models import (
    ClassPeriod,
    ClassPeriodSignUp,
    LibraryFacultyMember,
    Student,
    StudentInfo,
)


class CommonTestLogicMixin:
    """Contains common logic for testing spreadsheet-related functionality."""

    def setUp(self):
        # pylint: disable=invalid-name
        student1 = Student.objects.create_user(
            email="student1@myhchs.org", password="12345", name="Student1"
        )
        student2 = Student.objects.create_user(
            email="student2@myhchs.org", password="12345", name="Student2"
        )

        StudentInfo.objects.bulk_create(
            [
                StudentInfo(student=student1, id="123456"),
                StudentInfo(student=student2, id="654321"),
            ]
        )

        self.now = timezone.localtime(timezone.now())

        first_period = ClassPeriod(date=self.now, number=1, max_student_count=10)
        second_period = ClassPeriod(date=self.now, number=2, max_student_count=10)
        ClassPeriod.objects.bulk_create([first_period, second_period])

        # Creates ClassPeriodSignUp for both students.
        ClassPeriodSignUp.objects.bulk_create(
            [
                ClassPeriodSignUp(
                    student=student1,
                    class_period=first_period,
                    date_signed_up=self.now,
                    reason=ClassPeriodSignUp.LUNCH,
                    attendance_confirmed=True,
                    date_attendance_confirmed=self.now,
                ),
                ClassPeriodSignUp(
                    student=student2,
                    class_period=second_period,
                    date_signed_up=self.now,
                    reason=ClassPeriodSignUp.STUDY_HALL,
                    attendance_confirmed=True,
                    date_attendance_confirmed=self.now,
                ),
            ]
        )


class TestSpreadsheetGeneration(CommonTestLogicMixin, TestCase):
    """Tests :class:`signup.faculty.spreadsheets.generate_spreadsheet`."""

    def test_generate_spreadsheet(self):
        """Tests that the generated spreadsheet contains the correct headings and rows
        for the ClassPeriodSignUps."""
        workbook = generate_spreadsheet(ClassPeriodSignUp.objects.all())

        # Checks that the workbook only contains one sheet.
        self.assertEqual(len(workbook.worksheets), 1)

        sheet = workbook.active

        # Checks that there are only 3 rows: the headings and the two signups.
        rows = list(sheet.rows)
        self.assertEqual(len(rows), 3)

        # Checks that the first row contains the headings.
        first_row_values = [cell.value for cell in rows[0]]
        self.assertEqual(
            first_row_values,
            [
                "Date",
                "Period Number",
                "Student Name",
                "Student ID",
                "Reason",
                "Date Signed Up",
                "Attendance Confirmed",
                "Date Attendance Confirmed",
            ],
        )

        # Checks that the second row contains the first signup.
        second_row_values = [cell.value for cell in rows[1]]
        self.assertEqual(
            second_row_values,
            [
                self.now.date(),
                1,
                "Student1",
                "123456",
                "lunch",
                timezone.localtime(self.now).replace(tzinfo=None),
                True,
                timezone.localtime(self.now).replace(tzinfo=None),
            ],
        )

        # Checks that the third row contains the second signup.
        third_row_values = [cell.value for cell in rows[2]]
        self.assertEqual(
            third_row_values,
            [
                self.now.date(),
                2,
                "Student2",
                "654321",
                "study hall",
                timezone.localtime(self.now).replace(tzinfo=None),
                True,
                timezone.localtime(self.now).replace(tzinfo=None),
            ],
        )


class TestSpreadsheetView(CommonTestLogicMixin, TestCase):
    """Tests :class:`signup.faculty.views.GenerateSpreadsheetView`."""

    def setUp(self):
        super().setUp()

        library_faculty_member = LibraryFacultyMember.objects.create_user(
            email="faculty@myhchs.org", password="12345"
        )
        self.client.force_login(library_faculty_member)

    def test_generate_spreadsheet_full(self):
        """Tests that the HTTP response details are correct. Also tests that the
        generated spreadsheet contains the signups for a specific day."""
        # Checks that the HTTP headers are correct.
        response = self.client.get(
            reverse("signups_spreadsheet")
            + "?class_period__date="
            + self.now.strftime("%Y-%m-%d")
        )
        self.assertEqual(response["Content-Type"], "application/vnd.ms-excel")
        self.assertTrue("attachment" in response["Content-Disposition"])

        # Loads workbook from response.
        with BytesIO(response.content) as bytes_io:
            workbook = load_workbook(bytes_io)

        # Checks that the workbook only contains one sheet.
        self.assertEqual(len(workbook.worksheets), 1)

        sheet = workbook.active

        # Checks that there are only three rows: the headings and the two signups.
        rows = list(sheet.rows)
        self.assertEqual(len(rows), 3)

        # Checks that the first row contains the headings.
        first_row_values = [cell.value for cell in rows[0]]
        self.assertEqual(
            first_row_values,
            [
                "Date",
                "Period Number",
                "Student Name",
                "Student ID",
                "Reason",
                "Date Signed Up",
                "Attendance Confirmed",
                "Date Attendance Confirmed",
            ],
        )

        # Removes timezone information and microsecond precision since this data isn'r
        # present in Excel spreadsheets.
        now = timezone.localtime(self.now).replace(tzinfo=None, microsecond=0)

        second_row_values = [cell.value for cell in rows[1]]
        # Removes nanosecond and microsecond precision.
        second_row_values = [
            value.replace(second=int(value.second), microsecond=0)
            if isinstance(value, datetime)
            else value
            for value in second_row_values
        ]

        # Checks that the second row contains the first signup.
        self.assertEqual(
            second_row_values,
            [
                # Necessary since the corresponding value in second_row_values is a
                # datetime.datetime object without hours, minutes, or seconds.
                now.replace(hour=0, minute=0, second=0),
                1,
                "Student1",
                "123456",
                "lunch",
                now,
                True,
                now,
            ],
        )

        third_row_values = [cell.value for cell in rows[2]]
        # Removes nanosecond and microsecond precision.
        third_row_values = [
            value.replace(second=int(value.second), microsecond=0)
            if isinstance(value, datetime)
            else value
            for value in third_row_values
        ]

        # Checks that the third row contains the second signup.
        self.assertEqual(
            third_row_values,
            [
                # Necessary since the corresponding value in second_row_values is a
                # datetime.datetime object without hours, minutes, or seconds.
                now.replace(hour=0, minute=0, second=0),
                2,
                "Student2",
                "654321",
                "study hall",
                now,
                True,
                now,
            ],
        )

    def test_generate_spreadsheet_empty(self):
        """Tests that the HTTP response details are correct. Also tests that the
        generated spreadsheet will contain only one row (the headings) when there are no
        signups for a specific day."""
        # There are no ClassPeriodSignUps for tomorrow.
        tomorrow = self.now + timedelta(days=1)

        # Checks that the HTTP headers are correct.
        response = self.client.get(
            reverse("signups_spreadsheet")
            + "?class_period__date="
            + tomorrow.strftime("%Y-%m-%d")
        )
        self.assertEqual(response["Content-Type"], "application/vnd.ms-excel")
        self.assertTrue("attachment" in response["Content-Disposition"])

        # Loads workbook from response.
        with BytesIO(response.content) as bytes_io:
            workbook = load_workbook(bytes_io)

        # Checks that the workbook only contains one sheet.
        self.assertEqual(len(workbook.worksheets), 1)

        sheet = workbook.active

        # Checks that there is only one row: the headings.
        rows = list(sheet.rows)
        self.assertEqual(len(rows), 1)

        # Checks that the first row contains the headings.
        first_row_values = [cell.value for cell in rows[0]]
        self.assertEqual(
            first_row_values,
            [
                "Date",
                "Period Number",
                "Student Name",
                "Student ID",
                "Reason",
                "Date Signed Up",
                "Attendance Confirmed",
                "Date Attendance Confirmed",
            ],
        )
