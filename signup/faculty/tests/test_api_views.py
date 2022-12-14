from datetime import datetime, timedelta
from io import BytesIO

from django.core import mail
from django.test import TestCase
from django.urls import reverse
from django.utils import timezone
from openpyxl import load_workbook
from rest_framework.test import APIClient, APITestCase

from signup.faculty.tests.common import convert_datetime
from signup.faculty.tests.test_spreadsheets import (
    CommonTestLogicMixin as SpreadsheetTestLogicMixin,
)
from signup.models import (
    ClassPeriod,
    ClassPeriodSignUp,
    LibraryFacultyMember,
    Student,
    StudentInfo,
)


class CommonTestLogicMixin:
    """Contains common logic for testing views in :mod:`signup.faculty.api.views`."""

    def setUp(self):
        # pylint: disable=invalid-name
        self.student1 = Student.objects.create_user(
            email="student1@myhchs.org", password="12345", name="Student1"
        )
        StudentInfo.objects.create(student=self.student1, id="123456")

        self.student2 = Student.objects.create_user(
            email="student2@myhchs.org", password="12345", name="Student2"
        )
        StudentInfo.objects.create(student=self.student2, id="654321")

        self.library_faculty_member = LibraryFacultyMember.objects.create_user(
            email="faculty@myhchs.org", password="12345"
        )
        self.client.force_login(self.library_faculty_member)

        self.now = timezone.now()
        self.period = ClassPeriod.objects.create(
            date=self.now.date(), number=1, max_student_count=10
        )

        # Creates ClassPeriodSignUp for both students.
        self.signup1, self.signup2 = ClassPeriodSignUp.objects.bulk_create(
            [
                ClassPeriodSignUp(
                    id=1,
                    student=self.student1,
                    class_period=self.period,
                    date_signed_up=self.now,
                    reason=ClassPeriodSignUp.STUDY_HALL,
                ),
                ClassPeriodSignUp(
                    id=2,
                    student=self.student2,
                    class_period=self.period,
                    date_signed_up=self.now,
                    reason=ClassPeriodSignUp.STUDY_HALL,
                ),
            ]
        )


class TestClassPeriodSignUpListAPIView(CommonTestLogicMixin, APITestCase):
    """Performs tests on :class:`signup.faculty.views.ClassPeriodSignUpListAPIView`
    (excluding custom actions)."""

    def test_accessing_as_anonymous_user(self):
        """Tests that anonymous users will get an HTTP 401 error."""
        client = APIClient()

        response = client.get(reverse("api-signups-list"))
        self.assertEqual(response.status_code, 403)

    def test_accessing_as_student(self):
        """Tests that anonymous users will get an HTTP 403 error."""
        client = APIClient()
        client.force_login(self.student1)

        response = client.get(reverse("api-signups-list"))
        self.assertEqual(response.status_code, 403)

    def test_listing_periods_without_filtering(self):
        """Tests listing ClassPeriodSignUps by performing a GET request on
        ``api-signups-list``. Performs no filtering."""
        response = self.client.get(reverse("api-signups-list"))
        self.assertEqual(response.status_code, 200)
        self.assertListEqual(
            response.data,
            [
                {
                    "id": self.signup1.id,
                    "period_number": 1,
                    "student_name": "Student1",
                    "student_id": "123456",
                    "date_signed_up": convert_datetime(self.now),
                    "reason": ClassPeriodSignUp.STUDY_HALL,
                    "attendance_confirmed": False,
                    "date_attendance_confirmed": None,
                },
                {
                    "id": self.signup2.id,
                    "period_number": 1,
                    "student_name": "Student2",
                    "student_id": "654321",
                    "date_signed_up": convert_datetime(self.now),
                    "reason": ClassPeriodSignUp.STUDY_HALL,
                    "attendance_confirmed": False,
                    "date_attendance_confirmed": None,
                },
            ],
        )

    def test_listing_periods_with_filtering(self):
        """Tests listing ClassPeriodSignUps by performing a GET request on
        ``api-signups-list`` with url query parameters for filtering."""
        # Tests the DjangoFilterBackend filter backend.
        url = (
            reverse("api-signups-list")
            + f"?id={self.signup2.id}&class_period__number=1&student__name=Student2&student__info__id=654321"
        )
        response = self.client.get(url)
        self.assertEqual(response.status_code, 200)
        self.assertListEqual(
            response.data,
            [
                {
                    "id": self.signup2.id,
                    "period_number": 1,
                    "student_name": "Student2",
                    "student_id": "654321",
                    "date_signed_up": convert_datetime(self.now),
                    "reason": ClassPeriodSignUp.STUDY_HALL,
                    "attendance_confirmed": False,
                    "date_attendance_confirmed": None,
                }
            ],
        )

        # Tests the SearchFilter filter backend.
        url = reverse("api-signups-list") + "?search=Student2"
        response = self.client.get(url)
        self.assertEqual(response.status_code, 200)
        self.assertListEqual(
            response.data,
            [
                {
                    "id": self.signup2.id,
                    "period_number": 1,
                    "student_name": "Student2",
                    "student_id": "654321",
                    "date_signed_up": convert_datetime(self.now),
                    "reason": ClassPeriodSignUp.STUDY_HALL,
                    "attendance_confirmed": False,
                    "date_attendance_confirmed": None,
                }
            ],
        )

        # Tests multiple filter backends together (in this case, DjangoFilterBackend and
        # OrderingFilter).
        url = (
            reverse("api-signups-list")
            + "?class_period__number=1&ordering=-student__name"
        )
        response = self.client.get(url)
        self.assertEqual(response.status_code, 200)
        self.assertListEqual(
            response.data,
            [
                {
                    "id": self.signup2.id,
                    "period_number": 1,
                    "student_name": "Student2",
                    "student_id": "654321",
                    "date_signed_up": convert_datetime(self.now),
                    "reason": ClassPeriodSignUp.STUDY_HALL,
                    "attendance_confirmed": False,
                    "date_attendance_confirmed": None,
                },
                {
                    "id": self.signup1.id,
                    "period_number": 1,
                    "student_name": "Student1",
                    "student_id": "123456",
                    "date_signed_up": convert_datetime(self.now),
                    "reason": ClassPeriodSignUp.STUDY_HALL,
                    "attendance_confirmed": False,
                    "date_attendance_confirmed": None,
                },
            ],
        )


class TestClassPeriodSignUpDetailAPIView(CommonTestLogicMixin, APITestCase):
    """Performs tests on
    :class:`signup.faculty.views.ClassPeriodSignUpDetailAPIView`."""

    def test_accessing_as_anonymous_user(self):
        """Tests that anonymous users will get an HTTP 401 error."""
        client = APIClient()

        response = client.get(reverse("api-signups-detail", kwargs={"pk": "1"}))
        self.assertEqual(response.status_code, 403)

    def test_accessing_as_student(self):
        """Tests that anonymous users will get an HTTP 403 error."""
        client = APIClient()
        client.force_login(self.student1)

        response = client.get(reverse("api-signups-detail", kwargs={"pk": "1"}))
        self.assertEqual(response.status_code, 403)

    def test_retrieving_signup(self):
        """Tests getting info on a single ClassPeriodSignUp by performing a GET request
        on ``api-signups-detail``."""
        response = self.client.get(reverse("api-signups-detail", kwargs={"pk": "1"}))
        self.assertEqual(response.status_code, 200)
        self.assertDictEqual(
            response.data,
            {
                "id": self.signup1.id,
                "period_number": 1,
                "student_name": "Student1",
                "student_id": "123456",
                "date_signed_up": convert_datetime(self.now),
                "reason": ClassPeriodSignUp.STUDY_HALL,
                "attendance_confirmed": False,
                "date_attendance_confirmed": None,
            },
        )

    def test_updating_signup(self):
        """Tests updating info on a single ClassPeriodSignUp by performing a PATCH
        request on ``api-signups-detail``. Also tests that only writable fields are updated."""
        # Attempts to change both a writable field ("attendance_confirmed") and a
        # read-only field ("reason") for the first ClassPeriodSignUp.
        response = self.client.patch(
            reverse("api-signups-detail", kwargs={"pk": "1"}),
            {"attendance_confirmed": True, "reason": ClassPeriodSignUp.LUNCH},
        )

        # Checks that the request was valid and that only the writable field of
        # ClassPeriodSignUp ("attendance_confirmed") was modified.
        self.assertEqual(response.status_code, 200)
        self.assertDictEqual(
            response.data,
            {
                "id": self.signup1.id,
                "period_number": 1,
                "student_name": "Student1",
                "student_id": "123456",
                "date_signed_up": convert_datetime(self.now),
                "reason": ClassPeriodSignUp.STUDY_HALL,
                "attendance_confirmed": True,
                "date_attendance_confirmed": None,
            },
        )

    def test_deleting_signup_in_future(self):
        """Tests deleting a single ClassPeriodSignUp by performing a DELETE request on
        on ``api-signups-detail``. This ClassPeriodSignUp is associated with a period in
        the present/future. An email should be sent to the associated student in the
        process."""
        # Checks that the email outbox is empty.
        self.assertEqual(len(mail.outbox), 0)

        # Performs DELETE request.
        response = self.client.delete(reverse("api-signups-detail", kwargs={"pk": "1"}))
        self.assertEqual(response.status_code, 204)

        # Checks that an email has been sent to the student who signed up.
        self.assertEqual(len(mail.outbox), 1)
        self.assertEqual(mail.outbox[0].subject, "Media Center Sign-Up Removal")
        self.assertEqual(mail.outbox[0].to, ["student1@myhchs.org"])
        self.assertTrue(
            "You signed up to use the Holy Cross Media Center during period 1"
            in mail.outbox[0].body,
        )

        # Checks that the deleted ClassPeriodSignUp no longer exists.
        response = self.client.get(reverse("api-signups-detail", kwargs={"pk": "1"}))
        self.assertEqual(response.status_code, 404)

        # Checks that no more emails were sent.
        self.assertEqual(len(mail.outbox), 1)

    def test_deleting_signup_in_past(self):
        """Tests deleting a single ClassPeriodSignUp by performing a DELETE request on
        on ``api-signups-detail``. This ClassPeriodSignUp is associated with a period in
        the past. An email shouldn't be sent to the associated student."""
        past_period = ClassPeriod.objects.create(
            date=self.now - timedelta(days=3), number=1, max_student_count=10
        )
        ClassPeriodSignUp.objects.create(
            id=3,
            student=self.student1,
            class_period=past_period,
            date_signed_up=self.now,
            reason=ClassPeriodSignUp.STUDY_HALL,
        )

        # Checks that the email outbox is empty.
        self.assertEqual(len(mail.outbox), 0)

        # Performs DELETE request.
        response = self.client.delete(reverse("api-signups-detail", kwargs={"pk": "3"}))
        self.assertEqual(response.status_code, 204)

        # Checks that no email was sent to the student who signed up.
        self.assertEqual(len(mail.outbox), 0)

        # Checks that the deleted ClassPeriodSignUp no longer exists.
        response = self.client.get(reverse("api-signups-detail", kwargs={"pk": "3"}))
        self.assertEqual(response.status_code, 404)

        # Checks that no more emails were sent.
        self.assertEqual(len(mail.outbox), 0)

    def create_past_period(self):
        """Creates a period in the past."""
        # Creates a period in the past and a signups associated with that period.
        past_period = ClassPeriod.objects.create(
            date=self.now - timedelta(days=2), number=1, max_student_count=10
        )
        ClassPeriodSignUp.objects.create(
            id=3,
            student=self.student1,
            class_period=past_period,
            date_signed_up=self.now,
            reason=ClassPeriodSignUp.STUDY_HALL,
        )

    def test_deleting_multiple_signups_in_past(self):
        """Tests deleting multiple signups at once by performing a POST request on
        ``api-signups-delete-multiple``. Ensures that emails will not be sent to the
        students who signed up because the periods associated with the signups are in
        the past."""
        self.create_past_period()

        # Ensures that there are now three signups: two in the present/future and on in
        # the past.
        signups = ClassPeriodSignUp.objects.all()
        self.assertEqual(signups.count(), 3)

        # Performs POST request on signup associated with periods in the past.
        response = self.client.post(
            reverse("api-signups-delete-multiple"), {"id": ["3"]}
        )
        self.assertEqual(response.status_code, 204)

        # Ensures that only two signups remain.
        signups = ClassPeriodSignUp.objects.all()
        self.assertEqual(signups.count(), 2)

        # Ensures that no emails were sent.
        self.assertEqual(len(mail.outbox), 0)

    def test_deleting_multiple_signups_in_future(self):
        """Tests deleting multiple signups at once by performing a POST request on
        ``api-signups-delete-multiple``. Ensures that emails will be sent to the
        students who signed up because the periods associated with the signups are in
        the present/future."""
        self.create_past_period()

        # Performs POST request on signups associated with periods in the
        # present/future.
        response = self.client.post(
            reverse("api-signups-delete-multiple"),
            {"id": ["1", "2"]},
        )
        self.assertEqual(response.status_code, 204)

        # Ensures that only one signup remain.
        remaining = ClassPeriodSignUp.objects.all()
        self.assertEqual(remaining.count(), 1)

        # Ensures that two emails were sent because of the deleted signups.
        self.assertEqual(len(mail.outbox), 2)

        # Verifies the contents of the first email.
        self.assertEqual(mail.outbox[0].subject, "Media Center Sign-Up Removal")
        self.assertEqual(mail.outbox[0].to, ["student1@myhchs.org"])
        self.assertTrue(
            "You signed up to use the Holy Cross Media Center during period 1"
            in mail.outbox[0].body,
        )

        # Verifies the contents of the second email.
        self.assertEqual(mail.outbox[1].subject, "Media Center Sign-Up Removal")
        self.assertEqual(mail.outbox[1].to, ["student2@myhchs.org"])
        self.assertTrue(
            "You signed up to use the Holy Cross Media Center during period 1"
            in mail.outbox[1].body,
        )


class TestSpreadsheetView(SpreadsheetTestLogicMixin, TestCase):
    """Tests the custom action for generating a spreadsheet on
    :class:`signup.faculty.api.views.ClassPeriodSignUpViewSet`."""

    def setUp(self):
        super().setUp()

        library_faculty_member = LibraryFacultyMember.objects.create_user(
            email="faculty@myhchs.org", password="12345"
        )
        self.client.force_login(library_faculty_member)

    def test_generate_spreadsheet_full(self):
        """Tests that the HTTP response details are correct. Also tests that the
        generated spreadsheet contains the signups for a specific day."""
        # Checks that the HTTP headers are correct.
        response = self.client.get(
            reverse("api-signups-generate-spreadsheet")
            + "?class_period__date="
            + self.now.strftime("%Y-%m-%d")
        )
        self.assertEqual(response["Content-Type"], "application/vnd.ms-excel")
        self.assertTrue("attachment" in response["Content-Disposition"])

        # Loads workbook from response.
        with BytesIO(response.content) as bytes_io:
            workbook = load_workbook(bytes_io)

        # Checks that the workbook only contains one sheet.
        self.assertEqual(len(workbook.worksheets), 1)

        sheet = workbook.active

        # Checks that there are only three rows: the headings and the two signups.
        rows = list(sheet.rows)
        self.assertEqual(len(rows), 3)

        # Checks that the first row contains the headings.
        first_row_values = [cell.value for cell in rows[0]]
        self.assertEqual(
            first_row_values,
            [
                "Date",
                "Period Number",
                "Student Name",
                "Student ID",
                "Reason",
                "Date Signed Up",
                "Attendance Confirmed",
                "Date Attendance Confirmed",
            ],
        )

        # Removes timezone information and microsecond precision since this data isn'r
        # present in Excel spreadsheets.
        now = timezone.localtime(self.now).replace(tzinfo=None, microsecond=0)

        second_row_values = [cell.value for cell in rows[1]]
        # Removes nanosecond and microsecond precision.
        second_row_values = [
            value.replace(second=int(value.second), microsecond=0)
            if isinstance(value, datetime)
            else value
            for value in second_row_values
        ]

        # Checks that the second row contains the first signup.
        self.assertEqual(
            second_row_values,
            [
                # Necessary since the corresponding value in second_row_values is a
                # datetime.datetime object without hours, minutes, or seconds.
                now.replace(hour=0, minute=0, second=0),
                1,
                "Student1",
                "123456",
                "lunch",
                now,
                True,
                now,
            ],
        )

        third_row_values = [cell.value for cell in rows[2]]
        # Removes nanosecond and microsecond precision.
        third_row_values = [
            value.replace(second=int(value.second), microsecond=0)
            if isinstance(value, datetime)
            else value
            for value in third_row_values
        ]

        # Checks that the third row contains the second signup.
        self.assertEqual(
            third_row_values,
            [
                # Necessary since the corresponding value in second_row_values is a
                # datetime.datetime object without hours, minutes, or seconds.
                now.replace(hour=0, minute=0, second=0),
                2,
                "Student2",
                None,
                "study hall",
                now,
                True,
                now,
            ],
        )

    def test_generate_spreadsheet_empty(self):
        """Tests that the HTTP response details are correct. Also tests that the
        generated spreadsheet will contain only one row (the headings) when there are no
        signups for a specific day."""
        # There are no ClassPeriodSignUps for tomorrow.
        tomorrow = self.now + timedelta(days=2)

        # Checks that the HTTP headers are correct.
        response = self.client.get(
            reverse("api-signups-generate-spreadsheet")
            + "?class_period__date="
            + tomorrow.strftime("%Y-%m-%d")
        )
        self.assertEqual(response["Content-Type"], "application/vnd.ms-excel")
        self.assertTrue("attachment" in response["Content-Disposition"])

        # Loads workbook from response.
        with BytesIO(response.content) as bytes_io:
            workbook = load_workbook(bytes_io)

        # Checks that the workbook only contains one sheet.
        self.assertEqual(len(workbook.worksheets), 1)

        sheet = workbook.active

        # Checks that there is only one row: the headings.
        rows = list(sheet.rows)
        self.assertEqual(len(rows), 1)

        # Checks that the first row contains the headings.
        first_row_values = [cell.value for cell in rows[0]]
        self.assertEqual(
            first_row_values,
            [
                "Date",
                "Period Number",
                "Student Name",
                "Student ID",
                "Reason",
                "Date Signed Up",
                "Attendance Confirmed",
                "Date Attendance Confirmed",
            ],
        )
